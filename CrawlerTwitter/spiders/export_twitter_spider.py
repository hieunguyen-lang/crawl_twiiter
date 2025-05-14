# -*- coding: utf-8 -*-
__author__ = 'DuyLK'

import scrapy
import json
import random
import time
import sys
import redis
import mysql.connector
import openpyxl
from scrapy.conf import settings
from CrawlerTwitter.items import CrawlerTwitterItem
from scrapy.xlib.pydispatch import dispatcher
from scrapy import signals
from CrawlerTwitter.helper import HelperKeys
from helper import Helper
from scrapy.loader import ItemLoader
from datetime import datetime, timedelta
from scrapy.loader import ItemLoader
from mysql.connector import errorcode
from dispatcherLib import DispatcherLibrary
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.drawing.image import Image

reload(sys)
sys.setdefaultencoding('utf8')

class TwitterSpider(scrapy.Spider):
    name = "export_twitter"
    handle_httpstatus_list = [403, 404, 200, 401, 400]

    def __init__(self, lang='', company=0, maxResults=100, timeStart='', timeGet=5, *args, **kwargs):
        print('Init Spider ...')
        self.lang = lang
        self.company = int(company)
        self.maxResults = str(maxResults)
        self.timeGet = int(timeGet)
        self.countRequest = 0
        self.countNewItem = 0
        self.DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S"

        self.dataTwitter = []
        self.dataIDS = []
        self.dir_output = "/home/duylk/BigdataSupport/output/"
        self.keywords = [
            # "Bảo Việt Nhân thọ",

            # "Chubb Life",

            # "Dai-ichi Life",
            # "Dai-ichi",
            # "Daiichi",

            # "Cathay life",
            # "Cathay",

            # "Hanwha Life",

            # "Fubon Việt Nam",
            # "Fubon Life", 
            # "bảo hiểm Fubon", 
            # "Fubon Life",

            # "Generali",

            # "Phú Hưng life",

            # "Sun Life",

            # "BIDV Metlife",

            # "Mirae Asset Prévoir",
            # "MAP Life",

            # "Shinhan Life",

            # "Prudential",

            # "Manulife",

            # "AIA",

            # "FWD",

            "MB Ageas Life",
            "MBAL",

            # ---------------------------------------------------------------------------------------

            # "Home Credit",

            # "HD Saigon",
            # "HD Sài gòn",

            # "FE Credit",
            # "FE",

            # "Mirae Asset Finance Vietnam",
            # "Mirae Asset",
            # "Mirae",

            # "TIMA",

            # "Mcredit",

            # "Shinhan Finance",
            # "shinhan finance",
            # "sh finance",
            # "tín dụng shinhan",

            # "SHB Finance",
            # "shb finance",
            # "tín dụng shb",

            # "tín dụng",
            # "tăng trưởng tín dụng",
            # "chính sách tín dụng",
            # "tín dụng đen",

            # "cần vay",
            # "muốn vay",
            # "hỗ trợ vay",
            # "cho vay"
        ]

        #Redis Config
        self.redis_db = redis.Redis(host=settings['REDIS_HOST'], port=settings['REDIS_PORT'], db=settings['REDIS_DB_ID'])
        self.redis_Twitter_db = redis.Redis(host=settings['REDIS_HOST'], port=settings['REDIS_PORT'], db=settings['REDIS_DB_TWITTER_ID'])

        # Get Time
        if timeStart:
            self.timeStart = str(timeStart) + 'T23:59:59.000Z'
            timeFormatStart = datetime.strptime(self.timeStart, "%Y-%m-%dT%H:%M:%S.%fZ")
            self.timeEnd = str(timeFormatStart.date() - timedelta(days=self.timeGet)) + 'T00:00:00.000Z'
        else:
            self.timeStart = str(datetime.utcnow().date()) + 'T'+ str(datetime.utcnow().hour) +':'+ str(datetime.utcnow().minute - 1) + ':00.000Z'
            self.timeEnd = str(datetime.utcnow().date() - timedelta(days=self.timeGet)) + 'T00:00:00.000Z'

        print ("===========================")
        print (self.timeStart)
        print (self.timeEnd)
        print ("===========================")

        self.timeStart = '2021-09-21T23:59:59.000Z'
        self.timeEnd = '2021-09-18T00:00:00.000Z'

        print ("===========================")
        print (self.timeStart)
        print (self.timeEnd)
        print ("===========================")

        dispatcher.connect(self.spider_closed, signals.spider_closed)
        dispatcher.connect(self.spider_opened, signals.spider_opened)

    def start_requests(self):
        for keyword in self.keywords:
            headers = self.getHeaderRandom()
            print(keyword)
            if headers != '':
                if self.lang != '':
                    keyword = keyword+ " lang:"+self.lang
                # url = 'https://api.twitter.com/2/tweets/search/recent?query='+keyword+' lang:vi&start_time='+self.timeEnd+'&end_time='+self.timeStart+'&max_results='+self.maxResults+'&tweet.fields=public_metrics,author_id,created_at,entities,geo,in_reply_to_user_id,lang,possibly_sensitive,referenced_tweets,source'
                url = 'https://api.twitter.com/2/tweets/search/recent?query='+keyword+' lang:vi&max_results='+self.maxResults+'&tweet.fields=public_metrics,author_id,created_at,entities,geo,in_reply_to_user_id,lang,possibly_sensitive,referenced_tweets,source'
                yield scrapy.Request(url=url, headers=headers, meta={'keyword': keyword, 'headers': headers})
            else:
                print('[Error] Headers 1 -> Empty!')
                print('---------*0*----------')

    def parse(self, response):
        self.countRequest += 1
        #Response Meta
        posts = json.loads(response.body)
        # print('=======================================')
        # print posts
        # print('=======================================')
        metas = response.meta
        keyword = response.meta['keyword']
        headers = response.meta['headers']

        #API Error
        if response.status == 403 or response.status == 404:
            print ("[ERROR] 403 or 404 API: " + response.url)
            print (self.countHeader())
            if self.countHeader() > 0:
                headers = self.getHeaderRandom()
                if self.lang != '':
                    keyword = keyword+ " lang:"+self.lang
                url = 'https://api.twitter.com/2/tweets/search/recent?query='+keyword+' lang:vi&start_time='+self.timeEnd+'&end_time='+self.timeStart+'&max_results='+self.maxResults+'&tweet.fields=public_metrics,author_id,created_at,entities,geo,in_reply_to_user_id,lang,possibly_sensitive,referenced_tweets,source'
                print ("=>>>>>>>>>>>> Again Request to Twitter <<<<<<<<<<<<<<=")
                yield scrapy.Request(url=url, callback=self.parse, headers=headers, meta={'keyword': keyword})
            else:
                print('[Error] Headers 2 -> Empty!')
                # Helper().sendMessageTelegram('=>>>>>>>>>> Twitter Headers Key -> Empty')

        #API Empty
        if response.status == 400:
            print ("[ERROR] 400 API: " + response.url)

        if response.status == 401:
            print ("[ERROR] 401 API: Unauthorized")

        if response.status == 200:
            self.countNewItem += 1
            if "data" in posts:
                datas = posts['data']
                for post in datas:
                    post_id = post['id']
                    post_title = post['text']

                    datas = {"data": post, "meta": metas}
                    if self.countHeader() > 0:
                        headers = self.getHeaderRandom()
                        if self.lang != '':
                            keyword = keyword+ " lang:"+self.lang
                        url_user = "https://api.twitter.com/2/users/"+post['author_id']+"?user.fields=created_at,description,entities,id,location,name,pinned_tweet_id,profile_image_url,protected,public_metrics,url,username,verified,withheld"
                        yield scrapy.Request(url=url_user, callback=self.parseUser, headers=headers, meta=datas)
                    else:
                        print('[Error] Headers 3 -> Empty!')

            else:
                print('[ERROR] EMPTY RESURT!')
                print('---------*0*----------')
        else:
            print('[INFO] STATUS NOT 200')
            print('---------*0*----------')           

    def parseUser(self, response):
        print ("===")
        print ("[INFO] PARSER USER!")
        print ("===")
        users = json.loads(response.body)
        #API Empty
        if response.status == 400:
            print ("[ERROR] 400 API: " + response.url)

        if response.status == 401:
            print ("[ERROR] 401 API: Unauthorized")

        if response.status == 200:
            datas = response.meta
            if "data" in users:
                # item = self.detailPost(datas,users['data'])
                # if item != '':
                    # print item
                data = datas['data']
                metas = datas['meta']
                user = users['data']
                time = data['created_at']
                timeFormat = datetime.strptime(time, "%Y-%m-%dT%H:%M:%S.%fZ")
                timeCreated = (timeFormat + timedelta(hours=7)).strftime('%s')
                timeFormat =  (timeFormat + timedelta(hours=7)).strftime(self.DATETIME_FORMAT) 
                timeCreate = datetime.strptime(timeFormat, "%Y-%m-%d %H:%M:%S")
                url = "https://twitter.com/"+user['username']+"/status/"+data['id']

                if data['lang'] == "vi":
                    id_post = int(data['id'])
                    if id_post in self.dataIDS:
                        print ("------------------------------")
                        print ("[INFO] Duplicate")
                        print ("------------------------------")
                    else:
                        self.dataIDS.append(int(data['id'])) 
                        if timeCreate.month == 8 and timeCreate.year == 2021:
                            self.dataTwitter.append((
                                data['id'],
                                data['text'],
                                url,
                                data['author_id'],
                                user['name'],
                                user['username'],
                                user['profile_image_url'],
                                data['public_metrics']['like_count'],
                                data['public_metrics']['reply_count'],
                                data['public_metrics']['retweet_count'],
                                timeFormat
                            ))
                        else:
                            print (timeFormat)
                            print (url)
                            print ("[INFO] NOT IN MONTH == 8")
                else:
                    print ("===")
                    print ("[INFO] LANGUAGE: "+data['lang'])
                    print ("[INFO] NOT VIETNAMESE LANGUAGE")
                    print ("===")
                    return ''
                print('\n')
            else:
                print ("===")
                print ("[INFO] NOT REQUEST DATA USER!!!")
                print ("===")
        else:
            print ("===")
            print ("[INFO] STATUS NOT 200!!!")
            print ("===")

    #Get Key Random
    def getHeaderRandom(self):
        keys = HelperKeys.ApiKeyHelper.key
        if len(keys) > 0:
            key = keys[random.randint(0, len(keys) - 1)]
            return key
        else:
            return ''

    #Check Redis
    def redis_exists(self,key):
        val = self.redis_Twitter_db.get(key)
        if val == "exist":
            return True
        return False

    #Insert Key to Redis
    def insert_key_to_redis(self, key):
        nano = self.redis_Twitter_db.set(key,"exist")

        print ("================= SET =================")
        print ("=>>>>>>>>>>>> Key: " + key)
        print (nano)
        print ("=======================================")

    #Count Key
    def countHeader(self):
        keys = HelperKeys.ApiKeyHelper.key
        return len(keys)

    def spider_opened(self, spider):
        print ("Spider Open")

    def spider_closed(self, spider, reason):
        print("NUMBER REQUESTS: "+ str(self.countRequest))
        print("NUMBER OF NEW ITEMS: "+ str(self.countNewItem))
        print("DONE!")
        wb = openpyxl.load_workbook('/home/duylk/BigdataSupport/input/template/template_twitter.xlsx')
        sheet = wb[wb.sheetnames[0]]

        # START ROW EXCEL
        i = 3
        try:
            print ("[INFO] DATA TWITTER: "+str(len(self.dataTwitter))+"\n")
            # print self.dataTwitter
            count = 1
            for data in self.dataTwitter:
                idPost, description, urlPost, userID, name, username, userProfileImage, likeCount, replyCount, retweetCount, timeCreated = data
                
                keyword_search = ''
                flag = 0
                keyword_string = ''
                arr_keyword = []
                for keyword in self.keywords:
                    keyword_old = keyword
                    keyword = keyword.lower()
                    content = description.lower()
                    count_keyword_content = content.count(keyword)
                    count_totals = int(count_keyword_content)
                    if count_totals > 0:
                        keyword_search = keyword_old
                        arr_keyword.append(keyword_old)

                if len(arr_keyword) > 0:
                    keyword_string = ', '.join(str(e) for e in arr_keyword)
                    print (keyword_string)
                try:
                    if keyword_string:
                        print ("[INFO] ID POST: "+idPost)
                        sheet["A" + str(i)].value = "TWITTER"
                        sheet["B" + str(i)].value = name
                        sheet["C" + str(i)].value = timeCreated
                        sheet["D" + str(i)].value = description
                        sheet["E" + str(i)].value = urlPost
                        sheet["F" + str(i)].value = "POST"
                        sheet["G" + str(i)].value = likeCount
                        sheet["H" + str(i)].value = replyCount                
                        sheet["I" + str(i)].value = retweetCount
                        sheet["J" + str(i)].value = keyword_string
                        sheet["K" + str(i)].value = idPost
                        sheet["L" + str(i)].value = username
                        sheet["M" + str(i)].value = userID
                        i += 1
                    else:
                        print ("[INFO] NOT KEYWORD")
                except Exception as e:
                    print ("ERROR EXPORT: " + str(e))

            # SAVE TO FILE
            self.file_export = "Export_Youtube_"+self.keywords[0]+"_"+datetime.today().strftime("%d_%m_%Y")+"_V1.xlsx"
            self.file_name_export = self.dir_output+self.file_export
            wb.save(self.file_name_export)
            print("\n------------------------")
            print("SAVE FILE: " + self.file_name_export)
            print("\n------------------------")
        except Exception as e:
            print ("ERROR: " + str(e))